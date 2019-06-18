import os
import openpyxl
from openpyxl import load_workbook


# 将输出的内容保存到Excel
def writeExcel(dict, excel_file_path, filename='data.xlsx'):
    # addr = "D:\\work\\Excel\\test.xlsx"
    if not os.path.exists(excel_file_path):
        os.mkdir(excel_file_path)
    # 打开文件
    file_path = os.path.join(excel_file_path, filename)
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        # 创建一张新表
        wb = openpyxl.Workbook()  # 打开一个将写的文件
        ws = wb.create_sheet(index=0)  # 在将写的文件创建sheet
        ws.append(list(dict.keys()))
    ws.append(list(dict.values()))
    wb.save(file_path)  # 一定要记得保存
    print("数据保存到 %s" % file_path)


def writeExcel_tuple(tuple, excel_file_path, filename='data.xlsx'):
    if not os.path.exists(excel_file_path):
        os.mkdir(excel_file_path)
    # 打开文件
    file_path = os.path.join(excel_file_path, filename)
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        # 创建一张新表
        wb = openpyxl.Workbook()  # 打开一个将写的文件
        ws = wb.create_sheet(index=0)  # 在将写的文件创建sheet
    ws.append(tuple)
    wb.save(file_path)  # 一定要记得保存
    print("数据保存到 %s" % file_path)
# 提高效率，批量写入
def writeExcel_ext(data_list, excel_file_path):
    if len(data_list) < 1:
        return data_list
    if not os.path.exists(excel_file_path):
        os.mkdir(excel_file_path)
    # 打开文件
    file_path = os.path.join(excel_file_path, 'data_516.xlsx')
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        # 创建一张新表
        wb = openpyxl.Workbook()  # 打开一个将写的文件
        ws = wb.create_sheet(index=0)  # 在将写的文件创建sheet
        ws.append(list(data_list[0].keys()))
    i = 1
    for data in data_list:
        ws.append(list(data.values()))
        wb.save(file_path)  # 一定要记得保存
        print(str(i) + '...')
        i = i + 1

    print("数据保存到 %s" % file_path)